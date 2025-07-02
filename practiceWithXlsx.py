from openpyxl import load_workbook  # Allows reading and writing speadsheet files
import sys
import re


def main():
    ws = get_ws(sys.argv[1])
    APQL = get_APQL(ws)
    # num_of_samples = get_sample_num(ws)
    A_num = get_A_num(ws)
    swabs = determine_swabs(ws)
    the_list =  sample_set_org(determine_swabs(ws), get_swab_APQL(ws), get_A_num(ws), get_APQL(ws), get_sample_data(ws))
    print(sample_set_org(determine_swabs(ws), get_swab_APQL(ws), get_A_num(ws), get_APQL(ws), get_sample_data(ws)))
    # print(the_list['Limit'][0]['Material'].lower())
    statement_gen_rinse(sample_set_org(determine_swabs(ws), get_swab_APQL(ws), get_A_num(ws), get_APQL(ws), get_sample_data(ws)))


# TODO: Finish writing statment generator function
def statement_gen_rinse(sample_set_dict):
    if isinstance(sample_set_dict['Limit'], (int, float)):
        print('I\'m a rinse!')
    else:
        print('I\'m a swab!')
    for dicts in sample_set_dict['Sample List']:
# Permutation of all passing blank sample
        if (dicts['Sample Type'].lower() == 'blank' and
           dicts['Appearance'].lower() == 'pass' and
           isinstance(sample_set_dict['Limit'], (int, float)) and
           dicts['A-Result'] < sample_set_dict['Limit']):
            if isinstance(dicts['Other-Peak(s)'], (int, float)):
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Pass".
{sample_set_dict['Analyte']}: Peak not detected. Reported as "Pass".
Other Peak(s): {dicts['Other-Peak(s)']} ug/mL. Reported as "Pass".
Total({sample_set_dict['Analyte']} + Other Peak(s)): {dicts['Summed-Other-Peak(s)']} ug/mL.
''')
            elif isinstance(dicts['Other-Peak(s)'], str):
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Pass".
{sample_set_dict['Analyte']}: Peak not detected. Reported as "Pass".
Other Peak(s): {' ug/mL,'.join(dicts['Other-Peak(s)'].split(','))} ug/mL. Reported as "Pass".
Total({sample_set_dict['Analyte']} + Other Peak(s)): {dicts['Summed-Other-Peak(s)']} ug/mL.
''')
            else:
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Pass".
{sample_set_dict['Analyte']}: Peak not detected. Reported as "Pass".
Other Peak(s): Not detected. Reported as "Pass".
''')
# Permutation of all passing except appearance for blank rinse sample
        elif (dicts['Sample Type'].lower() == 'blank' and
           dicts['Appearance'].lower() == 'fail' and
           isinstance(sample_set_dict['Limit'], (int, float)) and
           dicts['A-Result'] < sample_set_dict['Limit']):
            if isinstance(dicts['Other-Peak(s)'], (int, float)):
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Fail".
{sample_set_dict['Analyte']}: Peak not detected. Reported as "Pass".
Other Peak(s): {dicts['Other-Peak(s)']} ug/mL. Reported as "Pass".
Total({sample_set_dict['Analyte']} + Other Peak(s)): {dicts['Summed-Other-Peak(s)']} ug/mL.
''')
            elif isinstance(dicts['Other-Peak(s)'], str):
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Fail".
{sample_set_dict['Analyte']}: Peak not detected. Reported as "Pass".
Other Peak(s): {' ug/mL,'.join(dicts['Other-Peak(s)'].split(','))} ug/mL. Reported as "Pass".
Total({sample_set_dict['Analyte']} + Other Peak(s)): {dicts['Summed-Other-Peak(s)']} ug/mL.
''')
            else:
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Fail".
{sample_set_dict['Analyte']}: Peak not detected. Reported as "Pass".
Other Peak(s): Not detected. Reported as "Pass".
''')
# Permutation of failing blank but passing appearance
        elif (dicts['Sample Type'].lower() == 'blank' and
           dicts['Appearance'].lower() == 'pass' and
           isinstance(sample_set_dict['Limit'], (int, float)) and
           dicts['A-Result'] > sample_set_dict['Limit']):
            if isinstance(dicts['Other-Peak(s)'], (int, float)):
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Pass".
{sample_set_dict['Analyte']}: Peak detected at {dicts['A-Result']} ug/mL. Reported as "Fail".
Other Peak(s): {dicts['Other-Peak(s)']} ug/mL. Reported as "Fail".
Total({sample_set_dict['Analyte']} + Other Peak(s)): {dicts['Summed-Other-Peak(s)']} ug/mL.
''')
            elif isinstance(dicts['Other-Peak(s)'], str):
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Pass".
{sample_set_dict['Analyte']}: Peak detected at {dicts['A-Result']} ug/mL. Reported as "Fail".
Other Peak(s): {' ug/mL,'.join(dicts['Other-Peak(s)'].split(','))} ug/mL. Reported as "Fail".
Total({sample_set_dict['Analyte']} + Other Peak(s)): {dicts['Summed-Total-Peak(s)']} ug/mL.
''')
            else:
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Pass".
{sample_set_dict['Analyte']}: Peak detected at {dicts['A-Result']} ug/mL. Reported as "Fail".
Other Peak(s): Not detected. Reported as "Pass".
Total({sample_set_dict['Analyte']} + Other Peak(s)): {dicts['A-Result']} ug/mL.
''')
# Permuation of failing appearance and A Result for blanks
        elif (dicts['Sample Type'].lower() == 'blank' and
           dicts['Appearance'].lower() == 'fail' and
           isinstance(sample_set_dict['Limit'], (int, float)) and
           dicts['A-Result'] > sample_set_dict['Limit']):
            if isinstance(dicts['Other-Peak(s)'], (int, float)):
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Fail".
{sample_set_dict['Analyte']}: Peak detected at {dicts['A-Result']} ug/mL. Reported as "Fail".
Other Peak(s): {dicts['Other-Peak(s)']} ug/mL. Reported as "Fail".
Total({sample_set_dict['Analyte']} + Other Peak(s)): {dicts['Summed-Other-Peak(s)']} ug/mL.
''')
            elif isinstance(dicts['Other-Peak(s)'], str):
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Fail".
{sample_set_dict['Analyte']}: Peak detected at {dicts['A-Result']} ug/mL. Reported as "Fail".
Other Peak(s): {' ug/mL,'.join(dicts['Other-Peak(s)'].split(','))} ug/mL. Reported as "Fail".
Total({sample_set_dict['Analyte']} + Other Peak(s)): {dicts['Summed-Total-Peak(s)']} ug/mL.
''')
            else:
                print(f'''
{dicts['Sample ID']} Blank
Appearance and Color: Clear and Colorless. Reported as "Fail".
{sample_set_dict['Analyte']}: Peak detected at {dicts['A-Result']} ug/mL. Reported as "Fail".
Other Peak(s): Not detected. Reported as "Fail".
Total({sample_set_dict['Analyte']} + Other Peak(s)): {dicts['A-Result']} ug/mL.
''')


# Makes a dict out of relevant sample set info
def sample_set_org(swab_state, swab_limits, analyte, limit, sample_list):
    sample_set_info = {}
    if swab_state:
        sample_set_info['Limit'] = swab_limits
    else:
        sample_set_info['Limit'] = limit
    sample_set_info['Analyte'] = analyte
    sample_set_info['Sample List'] = sample_list
    return sample_set_info


# Gather Swab APQL info return that info as a list of dicts
def get_swab_APQL(worksheet):
    material_APQLs = []
    for rows in worksheet.iter_rows(min_row=2, max_row=9, min_col=4, max_col=5, values_only=True):
        if rows[0] and isinstance(rows[1], (int, float)):
            material_APQLs.append({
                'Material': rows[0],
                'APQL': rows[1]
                })
        elif not rows[0] and not rows[1]:
            continue
        else:
            print(f'Please check the row of the {rows[0]} entry and that its inforamation is correct')
    return material_APQLs


# Gathers sample data and arranges it into a list of dicts, also automatically converts 'Other-Peak(s)' into one float
def get_sample_data(worksheet):
    sample_list = []
    for rows in worksheet.iter_rows(min_row=14, max_col=6, values_only=True):
        sample_list.append({
            'Sample ID': rows[0],
            'Sample Type': rows[1],
            'Material': rows[2],
            'Appearance': rows[3],
            'A-Result': rows[4],
            'Other-Peak(s)': rows[5],
            })
    for dicts in sample_list:
        if isinstance(dicts['Other-Peak(s)'], (int, float)):
            dicts['Summed-Other-Peak(s)'] = dicts['Other-Peak(s)']
            try:
                dicts['Summed-Total-Peak(s)'] = dicts['Other-Peak(s)'] + dicts['A-Result']
            except TypeError:
                sys.exit('An "A-Result" entry is non-numerical. Please fix and try again.')
        elif not dicts['Other-Peak(s)']:
            dicts['Summed-Other-Peak(s)'] = dicts['Other-Peak(s)']
            try:
                dicts['Summed-Total-Peak(s)'] = dicts['A-Result']
            except TypeError:
                sys.exit('An "A-Result" entry is non-numerical. Please fix and try again.')
        else:
            try:
                other_peaks = dicts['Other-Peak(s)'].split(',')
            except TypeError:
                sys.exit('One of more "Other-Peak(s)" entries are non-numerical values, please fix and try again')
            sum_of_peaks = 0
            for peaks in other_peaks:
                peaks = float(peaks)
                sum_of_peaks = sum_of_peaks + peaks
            sum_of_total_peaks = sum_of_peaks + float(dicts['A-Result'])
            dicts['Summed-Other-Peak(s)'] = sum_of_peaks
            dicts['Summed-Total-Peak(s)'] = sum_of_total_peaks
    return sample_list


# Gets the value of APQL
def get_APQL(worksheet):
    if isinstance(worksheet['B6'].value, float):
        return worksheet['B6'].value
    elif isinstance(worksheet['B6'].value, int):
        return worksheet['B6'].value
    else:
        sys.exit('APQL is not set to a numerical value. Please fix and try again.')


# Gets the number of samples
def get_sample_num(worksheet):
    try:
        if worksheet['B7'].value.is_integer():
            return worksheet['B7'].value
        else:
            sys.exit('Sample number is not set as an integer. Please fix and try again')
    except AttributeError:
        sys.exit('Sample number has non-numeric characters. Please fix and try again.')


# Gets name of target analyte
def get_A_num(worksheet):
    try:
        if worksheet['B5'].value:
            return worksheet['B5'].value
        else:
            sys.exit('Target Analyte is not set. Please fix and try again.')
    except IOError:
        sys.exit('Worksheet Error')


# Gets the active Worksheet from input Workbook
def get_ws(filename):
    if re.search('.+.xlsx+', filename):
        try:
            wb = load_workbook(filename=sys.argv[1])  # Spreadsheet should only have one sheet
            return wb.active
        except FileNotFoundError:
            sys.exit('File not Found!')
    else:
        sys.exit('File is not an ".xlsx" file')


# Determines if sample set are swabs or not
def determine_swabs(worksheet):
    try:
        if worksheet['B8'].value.strip().lower() == 'yes':
            return True
        else:
            return False
    except AttributeError:
        sys.exit('Swabs in Cell "B8" not set to yes or no. Please fix and try again')


main()
