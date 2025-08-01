from openpyxl import load_workbook  # Allows reading and writing speadsheet files
import sys
import re
from statement_gen_lib import state_gen


def main():
    if len(sys.argv) == 3:
        print('\n', state_gen(sample_list_checker(get_sample_data(get_ws(sys.argv[1])))), end='')
        with open(f'{sys.argv[2]}', 'w') as file:
            file.write(f'\n{state_gen(sample_list_checker(get_sample_data(get_ws(sys.argv[1]))))}')
    elif len(sys.argv) > 3:
        sys.exit('Please only link one file path at a time.')
    else:
        sys.exit('No file path given.')


# Gathers sample data and arranges it into a list of dicts, also automatically converts 'Other-Peak(s)' into one float
def get_sample_data(worksheet):
    sample_set_list = []
    for rows in worksheet.iter_rows(min_row=3, max_col=9, values_only=True):
        if (rows[0] and
                rows[1]):
            sample_set_list.append({
                'Sample ID': str(rows[0]).strip(),
                'Sample Type': rows[1].lower().strip(),
                'Units': rows[2].lower().strip(),
                'Limit': rows[3],
                'Appearance': rows[4].lower().strip(),
                'Analyte': rows[5].strip(),
                'Analyte RT': rows[6],
                'Analyte Result': rows[7],
                'Other-Peaks': rows[8],
                })
        else:
            continue
    return sample_set_list


def sample_list_checker(sample_list):
    sample_type_checker(sample_list)
    sample_unit_checker(sample_list)
    sample_limit_checker(sample_list)
    sample_appearance_checker(sample_list)
    sample_analyte_checker(sample_list)
    sample_ar_rt_checker(sample_list)
    sample_ar_cc_checker(sample_list)
    sample_other_checker(sample_list)
    return sample_list


def sample_type_checker(sample_list):
    for dicts in sample_list:
        if (dicts['Sample Type'] != 'blank' and
                dicts['Sample Type'] != 'sample'):
            raise SystemExit(sys.exit(f'Sample Type for {dicts['Sample ID']} is not of type "Blank" or "Sample"'))


def sample_unit_checker(sample_list):
    for dicts in sample_list:
        if (dicts['Units'] != 'ug/ml' and
                dicts['Units'] != 'ug/100cm2'):
            raise SystemExit(sys.exit(f'Units for {dicts['Sample ID']} is not either "ug/mL" or "ug/100cm2"'))


def sample_limit_checker(sample_list):
    for dicts in sample_list:
        if not isinstance(dicts['Limit'], (int, float)):
            raise SystemExit(sys.exit(f'Limit for {dicts['Sample ID']} is non-numerical'))


def sample_appearance_checker(sample_list):
    for dicts in sample_list:
        if (dicts['Appearance'] != 'pass' and
                dicts['Appearance'] != 'fail'):
            raise SystemExit(sys.exit(f'Appearance for {dicts['Sample ID']} is not either "Pass" or "Fail"'))


def sample_analyte_checker(sample_list):
    for dicts in sample_list:
        if not dicts['Analyte']:
            raise SystemExit(sys.exit(f'Not Analyte entry for {dicts['Sample ID']}, please fix and try again'))


def sample_ar_rt_checker(sample_list):
    for dicts in sample_list:
        if not isinstance(dicts['Analyte RT'], (int, float)):
            raise SystemExit(sys.exit(f'RT Entry for {dicts['Sample ID']} is non-numerical, please fix and try again'))


def sample_ar_cc_checker(sample_list):
    for dicts in sample_list:
        if not isinstance(dicts['Analyte Result'], (int, float)):
            raise SystemExit(sys.exit(f'Analyte Result for {dicts['Sample ID']} is non-numerical'))


def sample_other_checker(sample_list):
    for dicts in sample_list:
        if not dicts['Other-Peaks']:
            continue
        elif not isinstance(dicts['Other-Peaks'], (str)):
            raise SystemExit(sys.exit(f'Other-Peak(s) entrie(s) for {dicts['Sample ID']} not in proper format.'
                                      f'Please fix and try again.'))
        else:
            try:
                list_of_other_peaks = dicts['Other-Peaks'].split('), (')
            except ValueError:
                sys.exit(f'Other-Peak(s) entrie(s) for {dicts['Sample ID']} not in proper format.'
                         f'Please fix and try again.')
            else:
                list_of_other_peaks_dicts = []
                for pairs in list_of_other_peaks:
                    try:
                        _ = pairs.strip('()').split(',')
                        list_of_other_peaks_dicts.append({
                            'Retention Time': float(_[0].strip('()')),
                            'Concentration': float(_[1].strip('()'))
                            })
                        dicts['Other-Peaks'] = list_of_other_peaks_dicts
                    except ValueError:
                        sys.exit(f'Other-Peak(s) entrie(s) for {dicts['Sample ID']} not in proper format.'
                                 f'Please fix and try again.')


# Gets the active Worksheet from input Workbook
def get_ws(filename):
    if re.search('.+.xlsx+', filename):
        try:
            wb = load_workbook(filename=sys.argv[1])  # Spreadsheet should only have one sheet
            return wb.active
        except FileNotFoundError:
            sys.exit('File not found!')
    else:
        sys.exit('File is not an ".xlsx" file')


if __name__ == "__main__":
    main()
